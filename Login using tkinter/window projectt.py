from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib
root=Tk()
root.title("Data Entry")
root.configure(bg="#326273")
root.geometry("700x400+300+200")
root.resizable(False,False)
file=pathlib.Path("C:\\Users\\HP\\Desktop\\ex.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]=" name"
    sheet["B1"]="contact"
    sheet["C1"]="age"
    sheet["D1"]="gender"
    sheet["E1"]="address"
    file.save("C:\\Users\\HP\\Desktop\\ex.xlsx")
def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=ageValue.get()
    gender=gender_Combobox.get()
    address=addressEntry.get(1.0,END)
    file=openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\ex.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    file.save(r"C:\\Users\\HP\\Desktop\\ex.xlsx")
    messagebox.showinfo("info","detail saved")
def clear():
    nameValue.set("")
    contactValue.set("")
    ageValue.set("")
    addressEntry.delete(1.0,END)

    

    
#heading
Label(root,text="please fill out the entry form",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)
#labels
Label(root,text="Name",font=13,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="contact",font="arial 13",bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="age",font="arial 13",bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="gender",font="arial 13",bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="address",font="arial 13",bg="#326273",fg="#fff").place(x=50,y=250) 
nameValue=StringVar()
contactValue=StringVar()
ageValue=StringVar()
nameEntry=Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
nameEntry.place(x=200,y=100)
contactEntry=Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
contactEntry.place(x=200,y=150)
ageEntry=Entry(root,textvariable=ageValue,width=15,bd=2,font=20)
ageEntry.place(x=200,y=200)
gender_Combobox=Combobox(root,values=["male","female"],font='arial 13',state='r',width=14)
gender_Combobox.place(x=440,y=200)
gender_Combobox.set("male")
addressEntry=Text(root,width=50,height=4,bd=2)
addressEntry.place(x=200,y=250)
Button(root, text='Submit',width=20,bg='brown',fg='white',command=submit).place(x=200,y=350) 
Button(root, text='Clear',width=20,bg='brown',fg='white',command=clear).place(x=340,y=350)
Button(root, text='Exit',width=20,bg='brown',fg='white',command=lambda:root.destroy()).place(x=450,y=350)     
root.mainloop()